import asyncio
import re
from datetime import datetime
from http.cookies import BaseCookie

import aiohttp
from api_client import APIClient, APIFactory
from constants import BASE_URL, CHUNK_SIZE, EXCEL_FILE_PATH
from openpyxl import load_workbook
from openpyxl.worksheet.worksheet import Worksheet
from schemas import PatientBase
from schemas import PatientRecordsBase


class DataParser:

    @staticmethod
    async def parse_birthday(birthday: str) -> str:
        if isinstance(birthday, datetime):
            return birthday.date().isoformat()
        elif birthday in (None, "-", "нет даты"):
            return None
        else:
            try:
                return datetime.strptime(birthday, "%d.%m.%Y").date().isoformat()
            except (TypeError, ValueError):
                return None

    @staticmethod
    async def parse_living_place(living_place: str) -> str | None:
        return living_place.strip() if living_place else None

    @staticmethod
    async def parse_diagnosis(diagnosis: str | None) -> tuple[bool, bool, bool]:
        return (
            await DataParser.check_diagnosis(diagnosis, r"\bд[эе]п\b"),
            await DataParser.check_diagnosis(diagnosis, r"\bбп\b"),
            await DataParser.check_diagnosis(diagnosis, r"\bишемия\b"),
        )

    @staticmethod
    async def check_diagnosis(diagnosis: str | None, pattern: str) -> bool:
        return bool(re.search(pattern, diagnosis, flags=re.IGNORECASE)) if diagnosis else False

    @staticmethod
    async def determine_inhabited_locality(living_place: str) -> str:
        if living_place and living_place.strip().startswith(("г", "Г")):
            return "Город"
        return "Село" if living_place else "Неопределено"


class RowProcessor:

    @classmethod
    async def process_row(cls, row: tuple[str], cookies: BaseCookie, session: aiohttp.ClientSession) -> None:
        try:
            diagnosis = row[7]
            birthday = await DataParser.parse_birthday(row[2])
            living_place = await DataParser.parse_living_place(row[5])
            first_visit, last_visit = str(row[9]), str(row[8])
            treatment = row[10]
            dep, bp, ischemia = await DataParser.parse_diagnosis(diagnosis)
            inhabited_locality = await DataParser.determine_inhabited_locality(living_place)

            patient_data = PatientBase(
                full_name=row[0],
                birthday=birthday,
                gender=row[1],
                job_title=row[6],
                living_place=living_place,
                inhabited_locality=inhabited_locality,
                dep=dep,
                bp=bp,
                ischemia=ischemia,
            )
            response = await APIClient.create_patient(session, cookies, patient_data)
            patient_id = response["id"]

            first_patient_record_data = PatientRecordsBase(
                visit=first_visit, diagnosis=diagnosis, treatment=treatment, patient_id=patient_id
            )

            last_patient_record_data = PatientRecordsBase(
                visit=last_visit, diagnosis=diagnosis, treatment=treatment, patient_id=patient_id
            )

            await asyncio.gather(
                APIClient.create_patient_record(session, cookies, first_patient_record_data),
                APIClient.create_patient_record(session, cookies, last_patient_record_data),
            )

        except Exception as e:
            print(e)

    @classmethod
    async def parse_excel_sheet(
        cls, sheet: Worksheet, cookies: BaseCookie, session: aiohttp.ClientSession
    ) -> None:
        try:
            start_time = datetime.now()
            chunk = CHUNK_SIZE
            tasks = []
            pended = 0

            for _, row in enumerate(sheet.iter_rows(min_row=2, max_row=5337, values_only=True), start=1):
                tasks.append(asyncio.create_task(cls.process_row(row, cookies, session)))
                pended += 1
                if len(tasks) == chunk or pended == 5337:
                    print(pended)
                    await asyncio.gather(*tasks)
                    tasks = []

            end_time = datetime.now()
            print("Processing time:", end_time - start_time)

        except Exception as e:
            print(e)

    @classmethod
    async def authenticate_and_process(cls) -> None:
        async with await APIFactory.create_session() as session:
            cookies = await APIClient.authenticate(session)
            if not cookies:
                print({"message": "Failed to authenticate"})
            cookies = cls.get_cookies(session)
            wb = load_workbook(EXCEL_FILE_PATH)
            sheet = wb["Лист1"]
            await cls.parse_excel_sheet(sheet, cookies, session)
            wb.close()

    @staticmethod
    def get_cookies(session: aiohttp.ClientSession) -> BaseCookie:
        return session.cookie_jar.filter_cookies(BASE_URL)


class ExcelParser:

    @classmethod
    async def process_excel_file(cls) -> None:
        await RowProcessor.authenticate_and_process()
